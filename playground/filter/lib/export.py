"""Xuất kết quả đánh giá cascade ra file Excel."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from lib.cascade import FilterOutcome, FilterStats

_RESULT_COLUMNS = (
    ("result", "Result"),
    ("filter_stage", "Filter stage"),
    ("filter_reason", "Filter reason"),
    ("event_id", "Event ID"),
    ("source", "Source"),
    ("author_id", "Author"),
    ("timestamp", "Timestamp (UTC)"),
    ("likes", "Likes"),
    ("retweets", "Retweets"),
    ("followers", "Followers"),
    ("fasttext_label", "FastText label"),
    ("fasttext_score", "FastText score"),
    ("layers", "Layers"),
    ("clean_text", "Clean text"),
    ("raw_text", "Raw text"),
)


def default_export_path() -> Path:
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    return Path("exports") / f"filter_report_{ts}.xlsx"


def _ts_to_iso(ts: Any) -> str:
    if ts is None:
        return ""
    try:
        return datetime.fromtimestamp(int(ts), tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    except (TypeError, ValueError, OSError):
        return str(ts)


def _fasttext_meta(outcome: FilterOutcome) -> tuple[str, Any, list[str]]:
    meta = outcome.meta or {}
    layers: list[str] = []
    if isinstance(meta.get("layers"), list):
        layers = [str(x) for x in meta["layers"]]
    elif meta.get("layer"):
        layers = [str(meta["layer"])]

    ft = meta.get("fasttext") if isinstance(meta.get("fasttext"), dict) else {}
    if not isinstance(ft, dict):
        ft = meta if meta.get("label") or meta.get("score") is not None else {}

    label = str(ft.get("label") or "")
    score = ft.get("score")
    if score is not None:
        score = round(float(score), 4)
    return label, score, layers


def _row_from_pair(raw: dict[str, Any], outcome: FilterOutcome) -> dict[str, Any]:
    metrics = raw.get("metrics") or {}
    label, score, layers = _fasttext_meta(outcome)

    return {
        "result": "PASS" if outcome.passed else "DROP",
        "filter_stage": outcome.stage,
        "filter_reason": outcome.reason or "",
        "event_id": raw.get("event_id", ""),
        "source": raw.get("source", ""),
        "author_id": raw.get("author_id", ""),
        "timestamp": _ts_to_iso(raw.get("timestamp")),
        "likes": metrics.get("likes", ""),
        "retweets": metrics.get("retweets") or metrics.get("shares", ""),
        "followers": metrics.get("followers", ""),
        "fasttext_label": label,
        "fasttext_score": score,
        "layers": ", ".join(layers),
        "clean_text": outcome.clean_text,
        "raw_text": raw.get("raw_text", ""),
    }


def _write_sheet(
    wb: Workbook,
    title: str,
    rows: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet(title=title)
    header_font = Font(bold=True)
    keys = [key for key, _ in _RESULT_COLUMNS]

    for col, (_, label) in enumerate(_RESULT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(keys, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))

    if rows:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_RESULT_COLUMNS))}1"

    for col_idx, (key, label) in enumerate(_RESULT_COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        if key in ("raw_text", "clean_text"):
            ws.column_dimensions[letter].width = 60
        elif key == "event_id":
            ws.column_dimensions[letter].width = 38
        else:
            ws.column_dimensions[letter].width = min(max(len(label) + 2, 12), 40)


def _write_summary_sheet(
    wb: Workbook,
    stats: FilterStats,
    *,
    ml_available: bool,
    run_meta: dict[str, Any],
) -> None:
    ws = wb.active
    ws.title = "summary"
    bold = Font(bold=True)

    dropped = stats.dropped_l1 + stats.dropped_l2 + stats.dropped_l3
    pct_pass = (100.0 * stats.passed / stats.total) if stats.total else 0.0
    pct_drop = (100.0 * dropped / stats.total) if stats.total else 0.0

    ws["A1"] = "Filter evaluation report"
    ws["A1"].font = bold
    ws["A2"] = "Generated (UTC)"
    ws["B2"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")

    row = 4
    summary_rows = [
        ("Total processed", stats.total),
        ("PASS", f"{stats.passed} ({pct_pass:.1f}%)"),
        ("DROP", f"{dropped} ({pct_drop:.1f}%)"),
        ("", ""),
        ("DROP by layer", ""),
        ("  L1 heuristic", stats.dropped_l1),
        ("  L2 SimHash", stats.dropped_l2),
        ("  L3 FastText", stats.dropped_l3),
        ("", ""),
        ("FastText model", "available" if ml_available else "skipped / missing"),
    ]
    for label, value in summary_rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        if label in ("Filter evaluation report", "DROP by layer"):
            ws.cell(row=row, column=1).font = bold
        row += 1

    if stats.by_reason:
        row += 1
        ws.cell(row=row, column=1, value="DROP by reason").font = bold
        row += 1
        for key in sorted(stats.by_reason):
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=stats.by_reason[key])
            row += 1

    if run_meta:
        row += 1
        ws.cell(row=row, column=1, value="Run options").font = bold
        row += 1
        for key, value in run_meta.items():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40


def export_evaluation_report(
    *,
    passed: list[tuple[dict[str, Any], FilterOutcome]],
    dropped: list[tuple[dict[str, Any], FilterOutcome]],
    stats: FilterStats,
    path: Path,
    ml_available: bool,
    run_meta: dict[str, Any] | None = None,
) -> Path:
    """Xuất báo cáo đánh giá đầy đủ: summary + all + passed + dropped."""
    path = path.expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)

    all_rows = [_row_from_pair(raw, o) for raw, o in passed + dropped]
    passed_rows = [_row_from_pair(raw, o) for raw, o in passed]
    dropped_rows = [_row_from_pair(raw, o) for raw, o in dropped]

    wb = Workbook()
    _write_summary_sheet(wb, stats, ml_available=ml_available, run_meta=run_meta or {})
    _write_sheet(wb, "all", all_rows)
    _write_sheet(wb, "passed", passed_rows)
    _write_sheet(wb, "dropped", dropped_rows)

    wb.save(path)
    return path
